from openpyxl import load_workbook
from . import models 

def recupera_encabezados(hoja):
    encabezados = {}
    for encabezado in hoja.iter_rows(min_row=1, max_row=1, values_only=True):
        for index, value in enumerate(encabezado):
            encabezados[value] = index
    return encabezados 

def procesa_texto(texto):
    texto_formateado = texto.replace("\n", "<br>")
    return '<p><span style="color:#000000">' + texto_formateado + '</p>'

def get_or_create(model, **kwargs):
    instance, created = model.objects.get_or_create(**kwargs)
    return instance

# Catalogos Principales
def procesar_puntos(datos_fila):
    try:
        canal = models.CanalAcupuntura.objects.get( cvecanal=datos_fila['CveCanal'] )
        clave = datos_fila['CvePunto']
        chino = datos_fila['NomChinoPunto']
        nombre = datos_fila['NomPunto']
        models.PuntoAcupuntura.objects.get_or_create(cvecanal=canal, cvepunto=clave, nompunto=nombre, nomchino=chino)
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.CanalAcupuntura.DoesNotExist:
        print(f"No se encontró el canal con clave {datos_fila['CveCanal']} Para el Punto")

def procesar_canales(datos_fila):    
    models.CanalAcupuntura.objects.get_or_create( cvecanal=datos_fila['CveCanal'], nomcanal=datos_fila['NomCanal'])

def procesar_emociones(datos_fila):
    models.Emocion.objects.get_or_create( nombre=datos_fila['DescEmocion'])

def procesar_elementos(datos_fila):
    models.Elementos.objects.get_or_create( nombre=datos_fila['TipoPunto'])

def procesar_enfermedades(datos_fila):
    try:
        enfermedad = models.Enfermedad.objects.get( cveenfermedad=datos_fila['NomEnfermedad'])    
        enfermedad.nomenfermedad(datos_fila['NomEnfermedad'])
        enfermedad.save()
    except models.Enfermedad.DoesNotExist:
        models.Enfermedad.objects.get_or_create( cveenfermedad=datos_fila['NomEnfermedad'], nomenfermedad= datos_fila['NomEnfermedad'])

def procesar_sintomas(datos_fila):
    models.Sintoma.objects.get_or_create( sintoma=datos_fila['DescSintoma'])

# Puntos

def procesar_punto_significado(datos_fila):
    try:
        texto = procesa_texto(datos_fila['DescSignificado'])
        punto = models.PuntoAcupuntura.objects.get( cvepunto=datos_fila['CvePunto'] )
        significado = models.PuntoSignificado.objects.get(cvepunto=punto)
        significado.descsignificado(texto)
        significado.save()
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.PuntoAcupuntura.DoesNotExist:                
        print(f"No se encontró el punto con clave  {datos_fila['CvePunto']} para Punto Significado")
    except models.PuntoSignificado.DoesNotExist:
        if punto:
            models.PuntoSignificado.objects.get_or_create(cvepunto=punto, descsignificado=texto)

def procesar_punto_localizacion(datos_fila):
    try:
        texto = procesa_texto( datos_fila['DesLocalizacion'] )   
        punto = models.PuntoAcupuntura.objects.get( cvepunto=datos_fila['CvePunto'] )     
        localizacion = models.PuntoLocalizacion.objects.get(cvepunto=punto)
        localizacion.desclocalizacion(texto)
        localizacion.save()
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.PuntoAcupuntura.DoesNotExist:
        print(f"No se encontró el punto con clave {datos_fila['CvePunto']} para Punto Localizacion")
    except models.PuntoLocalizacion.DoesNotExist:
        if punto:
            models.PuntoLocalizacion.objects.get_or_create(cvepunto=punto, desclocalizacion=texto)

def procesar_punto_caracteristicas(datos_fila):
    try:
        punto = models.PuntoAcupuntura.objects.get( cvepunto=datos_fila['CvePunto'] )
        texto = procesa_texto( datos_fila['DescCaracteristica'] )
        models.PuntoCaracteristicas.objects.get_or_create(cvepunto=punto, desccaracteristicas=texto)
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.PuntoAcupuntura.DoesNotExist:
        print(f"No se encontró el punto con clave {datos_fila['CvePunto']} para Caracteristicas")

def procesar_punto_enfermedad(datos_fila):
    try:
        punto = models.PuntoAcupuntura.objects.get( cvepunto=datos_fila['CvePunto'] )
        enfermedad = models.Enfermedad.objects.get( nomenfermedad=datos_fila['NomEnfermedad'] )
        models.PuntoEnfermedad.objects.get_or_create(punto=punto, enfermedad=enfermedad)
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.PuntoAcupuntura.DoesNotExist:
        print(f"No se encontró el punto con clave {datos_fila['CvePunto']} para Punto Enfermedad")

def procesar_punto_sintomas(datos_fila):
    try:
        punto = models.PuntoAcupuntura.objects.get( cvepunto=datos_fila['CvePunto'] )        
        sintoma = models.Sintoma.objects.get( sintoma=datos_fila['DescSintoma'] )
        models.PuntoSintomatologia.objects.get_or_create(punto=punto, sintoma=sintoma)
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.PuntoAcupuntura.DoesNotExist:
        print(f"No se encontró el punto con clave {datos_fila['CvePunto']} para punto Sintoma")

# Canales


def procesar_canal_elemento(datos_fila):
    try:
        canal = models.CanalAcupuntura.objects.get( cvecanal=datos_fila['CveCanal'] )
        elemento = models.Elementos.objects.get( nombre=datos_fila['TipoPunto'] )
        models.CanalElemento.objects.get_or_create(canal=canal, elemento=elemento)
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.CanalAcupuntura.DoesNotExist:
        print(f"No se encontró el canal con clave {datos_fila['CveCanal']} Para Elemento")
    except models.Elementos.DoesNotExist:
        print(f"No se encontró el elemento con nombre {datos_fila['TipoPunto']}")


def procesar_canal_emocion(datos_fila):
    try:
        canal = models.CanalAcupuntura.objects.get( cvecanal=datos_fila['CveCanal'])
        emocion = models.Emocion.objects.get(nombre=datos_fila['DescEmocion'])
        models.CanalEmocion.objects.get_or_create(canal=canal, emocion=emocion)
    except KeyError as e:
        print(f"Falta un campo obligatorio en los datos de la fila: {e}")
    except models.CanalAcupuntura.DoesNotExist:
        print(f"No se encontró el canal con clave {datos_fila['CveCanal']} Para Emocion")
    except models.Emocion.DoesNotExist:
        print(f"No se encontró la emocion con nombre {datos_fila['DescEmocion']}")

# Tabla de elementos

def procesar_tabla_elementos(datos_fila):
    tipo_punto = get_or_create(models.Elementos, nombre=datos_fila['TipoPunto'])
    categoria = get_or_create(models.CategoriaElemento, nombre=datos_fila['Categoria'])
    valor = get_or_create(models.ValoresElemento, nombre=datos_fila['Valores'])
    models.TablaElemento.objects.get_or_create(tipo=tipo_punto, categoria=categoria, valor=valor)

#Procesar el archivo completo
 
def import_excel_file(excel_file, seleccionadas):
    print(f"Seleccionadas : {seleccionadas}")
    try:
        wb = load_workbook(filename=excel_file)
        #hoja = wb.active
        hojas_necesarias = ["ENFERMEDADES", "SINTOMAS", "CANALES", "EMOCIONES", "PUNTOS", 
                            "PUNTO_SIGNIFICADO", "PUNTO_LOCALIZACION", "PUNTO-CARACTERISTICAS", 
                            "PUNTO-ENFERMEDAD", "PUNTO-SINTOMATOLOGIA",
                            "CANAL_TIPOPUNTO", "CANAL-EMOCIONES", "PUNTO_TABLA_ELEMENTOS"]
        hojas_a_procesar = [hoja for hoja in hojas_necesarias if hoja in seleccionadas]
        todas_las_hojas_estan_presentes = all(nombre in wb.sheetnames for nombre in hojas_a_procesar)

        if todas_las_hojas_estan_presentes:
            for nombre_hoja in hojas_a_procesar:
                hoja = wb[nombre_hoja]
                columns = recupera_encabezados(hoja)

                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    datos_fila = {column_name: fila[column_index].strip() if isinstance(fila[column_index], str) else fila[column_index] for column_name, column_index in columns.items()}
                    try:
                        if nombre_hoja == "PUNTOS":
                            procesar_puntos(datos_fila)
                        elif nombre_hoja == "ENFERMEDADES":
                            procesar_enfermedades(datos_fila)
                        elif nombre_hoja == "SINTOMAS":
                            procesar_sintomas(datos_fila)
                        elif nombre_hoja == "CANALES":
                            procesar_canales(datos_fila)
                        elif nombre_hoja == "EMOCIONES":
                            procesar_emociones(datos_fila)
                        elif nombre_hoja == "TIPO-PUNTO": #Elementos                
                            procesar_elementos(datos_fila)

                        elif nombre_hoja == "PUNTO_SIGNIFICADO":                
                            procesar_punto_significado(datos_fila) # Relacion 1 a 1
                        elif nombre_hoja == "PUNTO_LOCALIZACION":              
                            procesar_punto_localizacion(datos_fila) # Relacion 1 a 1
                        elif nombre_hoja == "PUNTO-CARACTERISTICAS":              
                            procesar_punto_caracteristicas(datos_fila) # Relacion 1 a 1
                        elif nombre_hoja == "PUNTO-ENFERMEDAD":     # No repetir         
                            procesar_punto_enfermedad(datos_fila)
                        elif nombre_hoja == "PUNTO-SINTOMATOLOGIA": # No repetir
                            procesar_punto_sintomas(datos_fila) 
                        
                        elif nombre_hoja == "CANAL_TIPOPUNTO": # No repetir
                            procesar_canal_elemento(datos_fila)
                        elif nombre_hoja == "CANAL-EMOCIONES": # No repetir
                            procesar_canal_emocion(datos_fila)

                        elif nombre_hoja == "PUNTO_TABLA_ELEMENTOS":
                            procesar_tabla_elementos(datos_fila)
                    except Exception as e:
                        print(f"Fallo el procesamiento de :{nombre_hoja} - {fila} --- {e}" )    

            return f"Archivo procesado exitosamente las hojas: {hojas_a_procesar}"    
        else:
            return f"El archivo no tiene el formato adecuado para ser procesado {wb.sheetnames}"
    except:
        return f"No se pudo abrir el archivo {excel_file}"

